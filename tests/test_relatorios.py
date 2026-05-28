# =====================================================================
# tests/test_relatorios.py — Exportação .xlsx e coerência dos relatórios
# Verifica que o RH consegue baixar os relatórios em Excel e que os
# números mostrados batem com o cálculo (coerência entre relatórios).
# =====================================================================

from io import BytesIO
from datetime import datetime, timezone, date

import openpyxl

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _hoje_utc(h):
    t = date.today()
    return datetime(t.year, t.month, t.day, h, 0, tzinfo=timezone.utc)


def _textos_do_xlsx(conteudo):
    """Devolve o conjunto de todos os valores de célula (como str)."""
    wb = openpyxl.load_workbook(BytesIO(conteudo))
    textos = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    textos.add(str(v))
    return wb.sheetnames, textos


def test_frequencia_xlsx_baixa(app, org, login):
    """O relatório de frequência da equipe baixa como .xlsx válido."""
    c = app.test_client()
    login(c, org["rh_email"])
    resp = c.get("/rh/relatorios/frequencia.xlsx")
    assert resp.status_code == 200
    assert XLSX_MIME in resp.headers["Content-Type"]
    assert resp.data[:2] == b"PK"        # .xlsx é um zip → começa com 'PK'


def test_banco_horas_xlsx_baixa(app, org, login):
    """O banco de horas de um funcionário baixa como .xlsx válido."""
    c = app.test_client()
    login(c, org["rh_email"])
    resp = c.get(f"/rh/banco-horas/excel?user_id={org['func_id']}")
    assert resp.status_code == 200
    assert resp.data[:2] == b"PK"


def test_relatorios_sao_coerentes_com_o_calculo(app, org, login, criar_registro):
    """Coerência: 4h trabalhadas hoje devem aparecer como '04:00' tanto no
    banco de horas do funcionário quanto no relatório de frequência."""
    criar_registro(org["func_id"], "entrada", _hoje_utc(8),
                   lat=-20.34, lon=-40.29, local_id=org["local_id"])
    criar_registro(org["func_id"], "saida", _hoje_utc(12),
                   lat=-20.34, lon=-40.29, local_id=org["local_id"])

    hoje = date.today()
    c = app.test_client()
    login(c, org["rh_email"])

    # Banco de horas individual.
    bh = c.get(f"/rh/banco-horas/excel?user_id={org['func_id']}"
               f"&ano={hoje.year}&mes={hoje.month}")
    abas_bh, textos_bh = _textos_do_xlsx(bh.data)
    assert "Resumo" in abas_bh and "Detalhamento" in abas_bh
    assert "04:00" in textos_bh          # 4h trabalhadas

    # Frequência da equipe no mesmo período.
    freq = c.get(f"/rh/relatorios/frequencia.xlsx"
                 f"?periodo={hoje.year}-{hoje.month:02d}")
    _, textos_freq = _textos_do_xlsx(freq.data)
    assert "func" in textos_freq         # nome do funcionário
    assert "04:00" in textos_freq        # mesmo total → relatórios coerentes
